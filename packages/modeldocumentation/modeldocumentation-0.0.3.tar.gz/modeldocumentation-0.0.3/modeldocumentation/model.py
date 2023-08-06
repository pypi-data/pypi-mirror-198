import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.alignment import Alignment
import codecs
import json
import zipfile
import pathlib
from pathlib import Path
import openai
import os
from tkinter import filedialog, Tk
from tkinter.filedialog import askdirectory
import shutil
class model:
    def xls_extract_pbit(api_key,data,file,base_file_name,json_path):
        openai.api_key=api_key
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(title="Measures")
        cnt = 1
        for t in data['model']['tables']:
            if 'measures' in t:
                list_measures = t['measures']
                for i in list_measures:
                    prompt = "Explain the following calculation in a few sentences in simple business terms without using DAX function names: " + i['expression']
                    completion = openai.Completion.create(engine="text-davinci-003", prompt=prompt, max_tokens=64)
                    i['description'] = completion.choices[0]['text'].strip()
                    cnt += 1
                    if (cnt == 2): 
                        ws.append(['Measure Name', 'Measure Expression', 'Measure Data Type', 'Measure Description'])
                    ws.append([i['name'], i['expression'], i['dataType'] if 'dataType' in i else "", i['description']])
        if (cnt >= 2):
            table = Table(displayName="Table1", ref="A1:D" + str(cnt))
            ws.add_table(table)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            for col in ws.columns:
                for cell in col:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 30
            ws.column_dimensions["D"].width = 50
        else:
            ws.append(['No measures present in this file'])
            adjusted_width=(len('No measures present in this file')+2)*1.2
            ws.column_dimensions["A"].width=adjusted_width

        source = wb.create_sheet('Source Information')
        i = 0
        if 'tables' in data['model']:
            source.append(['Table No', 'Table Name', 'Table Mode', 'Table Type', 'Table Source','Original Table Name', 'Table Query', 'Modification','Modification Description'])
            for t in data['model']['tables']:
                if 'partitions' in t:
                    list_partitions = t['partitions']
                    List_source = (list_partitions[0]['source']['expression'])
                    expression=[]
                    if List_source[0] == 'let':
                        expression.append(List_source[0])
                        name = list_partitions[0]['name'].split('-')[0]
                        mode=str(list_partitions[0]['mode'])
                        i += 1
                        p = List_source[1]
                        expression.append(List_source[1])
                        Ttype = p.split("(")[0].split('= ')[1]

                        if '"' in p:
                            st = 0
                            ed = len(p) - 1
                            while (p[st] != '"'):
                                st += 1
                            while (p[ed] != '"'):
                                ed -= 1

                            TSource = p[st: ed + 1]
                            if "Query=" in TSource:
                                TSource = TSource.split("[Query")[0]
                            if "Delimiter=" in TSource:
                                TSource=TSource.split("),[Delimiter")[0]
                        else:
                            TSource = List_source[2]

                        otname = ""
                        if "Query=" in p:
                            otname=p.replace("#(lf)","")
                            otname=otname.split("FROM")[1]
                            if "WHERE" in otname: 
                                otname=otname.split("WHERE")[0]
                            if "INNER JOIN" in otname:
                                otname=otname.split("INNER JOIN")
                                n=[]
                                for tname in otname:
                                    if 'ON'in tname:
                                        n.append(tname.split("ON")[0])
                                    else:
                                        n.append(tname)
                                otname=",\n".join(n[0:])
                        elif "Sql." in p:
                            if "Source" in List_source[2]:
                                otname = List_source[3].split("=")[0].strip()
                            else:
                                otname = List_source[2].split("=")[0].strip()
                        elif "Excel." in p:
                            otname = List_source[2].split("=")[0].strip()
                            otname = otname.replace("#", "")
                        elif "Dataflows" in p:
                            otname = List_source[5].split("=")[0].split("\"")[1].strip()
                        else:
                            otname = name

                        TQuery = ""
                        if "Query=" in p:
                            TQuery = p.split("Query=")[1]
                            TQuery = TQuery.replace("#(lf)", " ")
                        else:
                            TQuery = "No Query"

                        idx = -1
                        for i1 in range(2, len(List_source)):
                            if "Sheet" not in List_source[i1]:
                                if len(List_source[i1]) > 5 and List_source[i1][4] == '#':
                                    idx = i1
                                    break
                        for i1 in range(2,idx):
                            expression.append(List_source[i1])
                        Tmodification = ""
                        Tdescription = ""
                        completion = "No Description"
                        if idx==-1:
                            Tmodification = "No Modification"
                            Tdescription = "No Description"
                        else:
                            pr = 1
                            for id in range(idx, len(List_source)-2):
                                p1 = List_source[id].split("    ")[1]
                                if '\\' not in p1:
                                    Tmodification += str(pr) + ". " + p1 + '\n\n'
                                    prompt= "Explain this in normal terms in one sentence: " + p1
                                    completion=openai.Completion.create(engine="text-davinci-003", prompt=prompt, max_tokens=64)
                                    to=completion.choices[0]['text'].strip().replace("\n","")
                                    Tdescription += str(pr) + ". " + to if completion != 'No Description' else completion
                                    Tdescription +=  '\n\n'
                                    to=completion.choices[0]['text'].strip().replace("\n","")
                                    to= "//"+to
                                    if completion != 'No Description':
                                        expression.append("    "+to)
                                    expression.append(List_source[id])
                                    pr += 1
                            expression.append(List_source[len(List_source)-2])
                            expression.append(List_source[len(List_source)-1])
                            t['partitions'][0]['source']['expression']=expression
                        source.append([i, name, mode, Ttype, TSource, otname, TQuery, Tmodification, Tdescription])

            with codecs.open(json_path, 'w', 'utf-16-le') as f:
                json.dump(data, f, indent=4)
        
            i += 1
            table = Table(displayName="Source", ref="A1:I" + str(i))
            source.add_table(table)

            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style

            for col in source.columns:
                for cell in col:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            source.column_dimensions["A"].width = 20
            source.column_dimensions["B"].width = 20
            source.column_dimensions["C"].width = 20
            source.column_dimensions["D"].width = 30
            source.column_dimensions["E"].width = 40
            source.column_dimensions["F"].width = 30
            source.column_dimensions["G"].width = 50
            source.column_dimensions["H"].width = 80
            source.column_dimensions["I"].width = 80
        else:
            source.append(['No Source present in this file'])
            adjusted_width=(len('No Source present in this file')+2)*1.2
            source.column_dimensions["A"].width=adjusted_width

        relation = wb.create_sheet('Relationships')
        if 'relationships' in data['model']:
            cnt1 = 0
            for t in data['model']['relationships']:
                direction = ""
                cardinality = ""
                if "joinOnDateBehavior" not in t:
                    cnt1 += 1
                    if "crossFilteringBehavior" in t:
                        direction = "Both Directional"
                    else:
                        direction = "Single Directional"

                    if "toCardinality" in t:
                        if t['toCardinality'] == "one":
                            cardinality = 'One to one (1:1)'
                        elif t['toCardinality'] == "many":
                            cardinality = 'Many to many (*:*)'
                        else:
                            pass

                    elif "fromCardinality" in t:
                        if t['fromCardinality'] == "one":
                            cardinality = 'One to one (1:1)'
                        elif t['fromCardinality'] == "many":
                            cardinality = 'Many to many (*:*)'
                        else:
                            pass
                    else:
                        cardinality = 'Many to one (*:1)'

                    if (cnt1 == 1): 
                        relation.append(['From Table', 'From Column', 'To Table', 'To Column', 'State', 'Direction', 'Cardinality'])
                    relation.append([t['fromTable'], t['fromColumn'], t['toTable'], t['toColumn'], t['state'] if 'state' in t else "", direction, cardinality])

            if (cnt1 >= 1):
                table = Table(displayName="Relationships", ref="A1:G" + str(cnt1 + 1))
                relation.add_table(table)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style

                for col in relation.columns:
                    for cell in col:
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                relation.column_dimensions["A"].width = 40
                relation.column_dimensions["B"].width = 20
                relation.column_dimensions["C"].width = 40
                relation.column_dimensions["D"].width = 20
                relation.column_dimensions["E"].width = 20
                relation.column_dimensions["F"].width = 20
                relation.column_dimensions["G"].width = 20

            else:
                relation.append(['No relationships present in this file'])
                adjusted_width=(len('No relationships present in this file')+2)*1.2
                relation.column_dimensions["A"].width=adjusted_width

        else:
            relation.append(['No relationships present in this file'])
            adjusted_width=(len('No relationships present in this file')+2)*1.2
            relation.column_dimensions["A"].width=adjusted_width

        dir_name = os.path.dirname(file)
        new_dir = pathlib.Path(dir_name, "EXCEL Output")
        new_dir.mkdir(parents=True, exist_ok=True)
        file_name = base_file_name + ".xlsx"
        save1 = str(new_dir) + "\\" + file_name
        wb.save(save1)
        wb = openpyxl.load_workbook(save1)
        sheet_to_remove = wb['Sheet']
        wb.remove(sheet_to_remove)
        wb.save(save1)
        print("Created :", file_name)


    def json_extract_pbit(api_key,file):
        temp_dir = 'temp'
        os.makedirs(temp_dir, exist_ok=True)
        dir_name = os.path.dirname(file)
        base_file_name = Path(file).stem

        pbit_path = Path(file)
        with zipfile.ZipFile(pbit_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        json_path = os.path.join(temp_dir, 'DataModelSchema')
        with codecs.open(json_path, 'r', 'utf-16-le') as f:
            contents = f.read()

        data = json.loads(contents)

        new_dir = pathlib.Path(dir_name, "JSON Output")
        new_dir.mkdir(parents=True, exist_ok=True)
        model.xls_extract_pbit(api_key,data, file, base_file_name, json_path)
        with codecs.open(json_path, 'r', 'utf-16-le') as f:
            contents = f.read()

        data = json.loads(contents)
        file_name = base_file_name + ".json"

        save1 = str(new_dir) + "\\" + file_name
        out_file = open(Path(save1), "w")

        json.dump(data, out_file, indent=6)
        out_file.close()
        print("Created :", file_name)

        dir_name = os.path.dirname(file)
        new_dir = pathlib.Path(dir_name, "UPDATED pbit")
        new_dir.mkdir(parents=True, exist_ok=True)

        new_pbit_path = base_file_name + '_updated.pbit'
        save1 = str(new_dir) + "\\" + new_pbit_path
        print("Created :", new_pbit_path)
        print()
        with zipfile.ZipFile(save1, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    rel_path = os.path.relpath(file_path, temp_dir)
                    zip_ref.write(file_path, rel_path)

        shutil.rmtree(temp_dir)

    def file(api_key,file):
        if(file != ""):
            if(file.endswith(".pbit")):
                model.json_extract_pbit(api_key,file)
            else:
                print("Enter a valid file")
        else:
            print("No file selected")

    def folder(api_key,folder):
        if(folder != ""):
            for f in os.listdir(folder):
                if f.endswith(".pbit"):
                    model.json_extract_pbit(api_key,f)
        else:
            print("No folder selected")

    def multi_files(api_key,files):
        if (files != ""):
            for f in files:
                if f.endswith(".pbit"):
                    model.json_extract_pbit(api_key,f)
        else:
            print("No files selected")
