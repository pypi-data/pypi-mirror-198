# -*- coding=utf-8 -*-
from IPython.display import display, HTML
from openpyxl import load_workbook


def render_table(ws, style=""):
    html = """
    <table style="{style}">
       <tr>{heading_row}</tr>
       {table_data}
    </table>
    """
    cell_names = " ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    heading_row = ""
    for i in range(ws.max_column+1):
        heading_row += "<th>{}</th>".format(cell_names[i])
    table_data = []
    for index, row in enumerate(ws.iter_rows()):
        row_data = "<tr><td>{}</td>{}</tr>".format(index, "<td>{}</td>"*ws.max_column)
        row_data = row_data.format(*[cell.value for cell in row])
        table_data.append(row_data)
    return html.format(style=style, heading_row=heading_row, table_data="".join(table_data))


def display_xlsx(excel_file, sheet_name, style="border:1px solid 666"):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    display(HTML(render_table(ws, style)))
