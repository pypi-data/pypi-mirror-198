
# ---------------------------------------------------------------------------------------------------------------
# PyxlSQL project
# This program and library is licenced under the European Union Public Licence v1.2 (see LICENCE)
# developed by fabien.battini@gmail.com
# ---------------------------------------------------------------------------------------------------------------

import re
import os
import importlib
import openpyxl
import openpyxl.styles
from PyxlSql.pyxlErrors import PyxlSqlSheetError, PyxlSqlError, PyxlSqlExecutionError
from PyxlSql.pyxlSheets import PyxlSheet

# ---------------------------------------------------------------------------------------------------
# class NamedWB
# ---------------------------------------------------------------------------------------------------


class PyxlWorkbook:
    """
    class to manage excel workbooks structured as a database
    """
    def __init__(self, file_name, create=False, column_name_row=1, font=None, file_path: list=[]):
        self.file_path = file_path or []
        self.filename = self.find_file(file_name)

        if self.filename is None:
            print(f"FATAL ERROR: Cannot find '{file_name}' in path '{self.file_path}', aborting")
            exit(-1)

        local_path = os.path.dirname(self.filename)
        if local_path not in self.file_path:
            self.file_path.append(local_path)

        try:
            self.wb: openpyxl.workbook = openpyxl.load_workbook(filename=self.filename)
        except OSError as error:
            print(f"FATAL ERROR: Cannot open '{file_name}' : {str(error)}, aborting")
            exit(-1)

        self.sheets = {}  # a dictionary  string --> NamedWS
        self.book_column_name_row = column_name_row  # BookColumnNameRow  column_name_row
        self.book_default_font = openpyxl.styles.Font(name='Century Gothic', size=11) if font is None else font
        self.wbs = {}       # all workbooks referenced by this one
        self.imported = {}  # dictionary of imported symbols, will be used when eval is called
        self.sheets_to_delete = []
        self.import_module("functools")
        if create:
            self.wb = openpyxl.Workbook()
            return

        for sheet in self.wb:
            self.sheets[sheet.title] = PyxlSheet(self, sheet, column_name_row=column_name_row)

    def find_file(self, filename):
        if filename is None:
            return None
        if os.path.exists(filename):
            return filename
        for d in self.file_path:
            if d[-1] != "/":
                d += "/"
            f = os.path.realpath(d + filename)
            if os.path.exists(f):
                return f
        return None

    def get_sheet(self, sheet_name, raise_exception=True):
        if sheet_name is None:
            return None
        if sheet_name in self.sheets:
            return self.sheets[sheet_name]

        m = re.search("\"?\'?([^\"\'[]+)\"?\'?(\\[(.*)])?", sheet_name)
        if m is not None:
            first_name: str = m.group(1)
            second_name: str = m.group(3)
            if first_name is not None and second_name is not None:
                workbook = self.get_workbook(first_name)
                if workbook is not None and second_name in workbook.sheets:
                    return workbook.sheets[second_name]
            # here, we have not found the sheet

        if raise_exception:
            raise PyxlSqlSheetError(sheet_name, "workbook")
        return None

    def delete_sheet(self, sheet_name):
        """Deletes the sheet from the workbook"""
        sheet = self.get_sheet(sheet_name)
        if sheet is not None:
            del self.wb[sheet_name]

    def rename_sheet(self, old_sheet_name:str, new_sheet_name:str):
        """Rename the sheet from the workbook"""
        sheet = self.get_sheet(old_sheet_name)
        # sheet cannot be None, get_sheet would have raised an exception

        sheet.rename(new_sheet_name)
        self.sheets[new_sheet_name]=sheet
        self.sheets.pop(old_sheet_name)

    def save(self, file_name):
        current_dir = os.path.dirname(os.path.realpath(self.filename))
        try:
            self.wb.save(current_dir + "\\" + file_name)
        except OSError as err:
            raise PyxlSqlExecutionError(f" file '{current_dir}\\{file_name}' is read-only", str(err))
    def get_workbook(self, name):
        if name not in self.wbs:
            self.wbs[name] = PyxlWorkbook(name, file_path=self.file_path)
        return self.wbs[name]

    def import_module(self, module, sub_modules=None):
        mod = importlib.import_module(module)
        if sub_modules is None:
            self.imported[module] = mod
            return

        if sub_modules == '*':
            if hasattr(mod, '__all__'):
                item_list = mod.__all__
            else:
                raise PyxlSqlError(f"ERROR: IMPORT {module} SUBS {sub_modules} : ABORTING the import",
                                   "        The module does not contain a __all__ symbol that allows importing *")
        else:
            item_list = sub_modules.split()

        for item in item_list:
            self.imported[item] = getattr(mod, item)
