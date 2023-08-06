
# ---------------------------------------------------------------------------------------------------------------
# PyxlSQL project
# This program and library is licenced under the European Union Public Licence v1.2 (see LICENCE)
# developed by fabien.battini@gmail.com
# ---------------------------------------------------------------------------------------------------------------

import re
import inspect
import openpyxl
import openpyxl.styles
from PyxlSql.pyxlErrors import PyxlSqlColumnError


# ---------------------------------------------------------------------------------------------------------------
# class NamedWS
# ---------------------------------------------------------------------------------------------------------------


class PyxlSheet:
    """
    Additional metadata for each named sheet in the NamedWB
    NB: Column names are *case-sensitive* (NOT converted to lower case)
    """
    general_style = 'General'
    int_style = '#,##0'
    percent_style = '0.0%'
    euro_style = '_ * #,##0_) [$€-x-euro1]_ '
    euro2_style = '_ * #,##0.00_) [$€-x-euro1]_ '

    def __init__(self, workbook, sheet, column_name_row=1, font=None):
        sheet.auto_filter.ref = None    # to avoid issues when auto_filter is not correctly managed by OpenPyxl

        self.openpyxl_sheet = sheet
        self.father = workbook

        self.full_name = self.father.filename + ':[' + sheet.title + ']'
        self.current_row = sheet.max_row + 1    # first free row
        self.column_name_row = column_name_row  # number of the row where the column name is stored
        self.columns = []                       # a list of columns
        self.column_names = {}                  # a dictionary: string --> Column number
        self.column_styles = {}                 # a dictionary: string --> style
        self.indexes = {}                       # a dictionary: index -->
        self.book_default_font = openpyxl.styles.Font(name='Century Gothic', size=11) if font is None else font

        max_col = sheet.max_column+1
        for i in range(1, max_col):
            col_name = sheet.cell(row=column_name_row, column=i).value
            if col_name is not None and col_name != "" and col_name != " ":
                col_name = str(col_name)  # str(), because could be an integer, typically a year .lower()
                self.column_names[col_name] = i
                self.columns.append(col_name)
                cell = sheet.cell(row=column_name_row + 1, column=i)
                if cell is not None:
                    self.column_styles[col_name] = cell.number_format

    def rename(self, new_name:str):
        self.openpyxl_sheet.title = new_name
        self.full_name = self.father.filename + ':[' + new_name + ']'

    def get_row_range(self):
        """returns the range of ACTIVE rows"""
        return range(self.column_name_row + 1, self.current_row)

    def get_start_of_range(self):
        return self.column_name_row + 1

    def find_column(self, column_name):
        """returns the column number from its name"""
        if column_name is None:
            raise PyxlSqlColumnError("find_column(None)", self.full_name)

        lc_column_name = str(column_name)  # .lower()
        if lc_column_name not in self.column_names:
            raise PyxlSqlColumnError(column_name, self.full_name)
        return self.column_names[lc_column_name]

    def get_max_column(self):
        # Returns the ID of the last column
        return len(self.column_names)+1

    def get_column_range(self):
        return range(1, self.get_max_column())

    @staticmethod
    def build_key(*args):
        """
        Creates the hash key for a list of descriptors,
        Typically (Column name, Row name)

        CAVEAT: all identifiers are NO MORE turned into lower case
        """
        key = ""
        for v in args:
            key += "" if v is None else (":" + str(v))  # str(v).lower())
        return key

    def get_cell(self, row_nb, column_name):
        col = self.find_column(column_name)
        if row_nb is None:
            raise PyxlSqlColumnError("Undefined",
                                     f"source:{str(inspect.currentframe().f_lineno)}" +
                                     f":{str(inspect.currentframe().f_code.co_filename)}")
        return self.openpyxl_sheet.cell(row=row_nb, column=col)

    def get_val_by_nb(self, row_nb, col):
        cell = self.openpyxl_sheet.cell(row=row_nb, column=col)
        return None if cell.value is None or cell.value == "" else cell.value

    def get_val(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        return None if cell.value is None or cell.value == "" else cell.value

    # UNUSED, for future PIVOT
    def get_string(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        if cell is None or cell.value is None or cell.value == "None":
            return ""
        return str(cell.value)

    @staticmethod
    def get_cell_color(cell):
        my_color = cell.fill.fgColor.rgb
        if isinstance(my_color, str):
            return my_color[2:]  # skip Alpha
        return None

    def get_float(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        if cell is None:
            return 0
        val = cell.value

        if val is None or val == "":
            return 0

        if isinstance(val, str) and val.startswith("="):
            val = val[1:]  # remove leading =
            int_val = eval(val)  # a tentative to manage simple cells, such as '=12.56-24.9'
            # print("EVAL", val, intVal)
            # TODO: Track Errors
            return int_val

        return float(val)  # ********** CAVEAT! must be Float, because can be less than 1

    def get_style(self, column_name):
         lc_column_name = str(column_name)  # .lower()
         if lc_column_name not in self.column_names:
             raise PyxlSqlColumnError(column_name, self.full_name)
         return self.column_styles[lc_column_name]

    # def get_index(self, *column_names):
    #     """
    #     :param self: a sheet
    #     :param column_names: the list of columns to be indexed
    #     :return: the index, i.e. the dictionary (value of the indexed columns) --> row number
    #
    #     if the index was not yet created, creates it
    #     """
    #     key = NamedWS.build_key(*column_names)  # Manages SEVERAL columns
    #     if key not in self.indexes:
    #         index_hash = {}
    #         for row in self.get_row_range():
    #             values = []
    #             for col in column_names:
    #                 values.append(self.get_val(row, col))
    #             val_key = NamedWS.build_key(*values)
    #             if val_key not in index_hash:
    #                 index_hash[val_key] = []
    #             index_hash[val_key].append(row)
    #         self.indexes[key] = index_hash
    #     return self.indexes[key]

    # ------------ set methods

    def set_value(self, row_nb: int, column_name: str, value, number_format=None) -> None:
        cell = self.get_cell(row_nb, column_name)
        assert cell is not None, "INTERNAL: set_value"

        if isinstance(value, str):
            value = re.sub("^ *", "", value)
            value = re.sub(" *$", "", value)

        cell.value = value
        cell.font = self.book_default_font
        style = number_format if number_format is not None else self.get_style(column_name)
        if style is not None:
            cell.number_format = style

        if number_format:
            cell.number_format = number_format

        self.current_row = max(self.current_row, row_nb+1)  # current_row is the first FREE cell

    # UNUSED, for future PIVOT
    # def create_row(self, init):
    #     for column, value in init.items():
    #         self.set_value(self.current_row, column, value)
    #
    #     self.current_row += 1
    #     return self.current_row - 1

    @staticmethod
    def get_col_from_int(column_nb):
        column_nb -= 1
        my_max = ord('Z') - ord('A') + 1
        low = column_nb % my_max
        res = chr(ord('A') + low)
        high = int((column_nb - low) / my_max)

        if high > 0:
            res = chr(ord('A')+high-1) + res
        return res
    def get_reference(self, row_nb, column_name, absolute=True, same_sheet=True):
        if same_sheet:
            sheet_ref = ""
        else:
            sheet_ref = "'" + self.openpyxl_sheet.title + "'!"

        if absolute:
            ref = "{}${}${}".format(sheet_ref, self.get_col_from_int(self.find_column(column_name)),
                                    row_nb)
        else:
            ref = "{}{}{}".format(sheet_ref, self.get_col_from_int(self.find_column(column_name)),
                                  row_nb)
        return ref